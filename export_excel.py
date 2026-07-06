#!/usr/bin/env python3
"""Export all hidden targets to Excel with PubMed validation status."""
import torch, json, time, math, os
from collections import defaultdict, deque
from datetime import datetime
import urllib.request, urllib.parse, xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def log(msg):
    print(msg, flush=True)

t0 = time.time()
log("=" * 60)
log("HIDDEN TARGETS EXCEL EXPORT")
log(f"{datetime.now():%H:%M:%S}")
log("=" * 60)

# ── Load model + data ──
log("\n[1/4] Loading PCA model + computing hidden targets...")
data = torch.load("data/processed/heterodata.pt", weights_only=False)
feats = torch.load("checkpoints/pca_features/features_128d.pt", weights_only=False)
for nt in data.node_types:
    data[nt].x = feats[nt]

ets = [(et, data[et].edge_index.shape[1]) for et in data.edge_types if et[1] != 'MENTIONED_IN']
ets.sort(key=lambda x: -x[1])
top20 = [et for et, n in ets[:20]]
temp_init = {f'{s}__{r}__{d}': 1.0 for s, r, d in top20}

from models.tempered_hgt import TemperedHGT
model = TemperedHGT(
    in_channels={nt: 128 for nt in data.node_types},
    hidden_channels=128, out_channels=128, num_heads=4, num_layers=2,
    meta_relations=top20, temperature_init=temp_init,
)
with open("checkpoints/pca_features/summary.json") as f:
    summary = json.load(f)
ckpt = torch.load(summary["checkpoint"], weights_only=True)
model.load_state_dict(ckpt["model_state_dict"])
model.eval()

target_ets = [et for et in top20 if et[2] == 'Disease' and et[0] in ('Gene', 'Protein', 'Drug', 'Cytokine')]
ei_dict = {et: data[et].edge_index for et in top20}
x_dict = {nt: data[nt].x for nt in data.node_types}
with torch.no_grad():
    z = model(x_dict, ei_dict, cos_decay=0.0)

predictions = []
for et in target_ets:
    ei = data[et].edge_index
    with torch.no_grad():
        scores = model.decode(z, ei, et[0], et[2])
    sn = data[et[0]].name if hasattr(data[et[0]], 'name') else []
    for i in range(ei.shape[1]):
        predictions.append({
            'name': sn[int(ei[0,i])] if int(ei[0,i]) < len(sn) else str(int(ei[0,i])),
            'src_type': et[0], 'relation': et[1],
            'score': float(scores[i]), 'src_idx': int(ei[0,i]),
        })

# Blacklist
BL = {'f2','prothrombin','thrombin','f5','f7','f8','f9','f10','f11','f12','f13',
      'tf','tissue factor','vwd','vwf','von willebrand','adamts13','fxa','factor xa',
      'fva','fvii','fviii','fix','fxi','fxii','protein c','protein s','antithrombin',
      'serpine1','pai-1','tfp1','thrombomodulin','epcr','fibrinogen','fibrin','fga',
      'fgb','fgg','plasminogen','plg','tpa','plasmin','p-selectin','selectin',
      'gpiib','gpiia','gpib','gpvi','il-6','il-1','il-8','tnf','nf-kb','nfkb','tlr4',
      'tlr2','cox-2','il-10','il-12','il-4','vegf','nos3','eno','endothelin',
      'factor v','mthfr','jak2','hmgb1','padi4','albumin','glucose','cholesterol',
      'triglyceride','renin','c3','mmp-9','tsp-1','inos','bdnf','tlr5','par-2'}
ANCHORS = {'fut8','lgals3','galectin-3','cd44','itgb1','itga5','itgav','mapk1',
           'mapk3','rhoa','rock1','rock2','nfkb1','rela'}

def is_bl(n):
    for t in BL:
        if t in n.lower(): return True
    return False

def is_anchor(n):
    for a in ANCHORS:
        if a in n.lower(): return True
    return False

# Degree + adjacency
degree = defaultdict(int)
for et in data.edge_types:
    ei = data[et].edge_index
    for j in range(ei.shape[1]):
        degree[(et[0], int(ei[0,j]))] += 1
        degree[(et[2], int(ei[1,j]))] += 1

adj = defaultdict(list)
for et in data.edge_types:
    ei = data[et].edge_index
    for j in range(ei.shape[1]):
        s, d = int(ei[0,j]), int(ei[1,j])
        adj[(et[0],s)].append((et[2],d,et[1]))
        adj[(et[2],d)].append((et[0],s,f'rev_{et[1]}'))

# Compute hidden targets
hidden = []
for p in predictions:
    if is_bl(p['name']) or is_anchor(p['name']):
        continue
    d = degree.get((p['src_type'], p['src_idx']), 1)
    ds = p['score'] / math.log(d + 1)
    # BFS to anchors
    start = (p['src_type'], p['src_idx'])
    visited = {start: 0}
    q = deque([start])
    ah = None
    while q and ah is None:
        cur = q.popleft()
        dist = visited[cur]
        if dist >= 3:
            continue
        for nbr in adj.get(cur, []):
            nk = (nbr[0], nbr[1])
            if nk in visited:
                continue
            visited[nk] = dist + 1
            q.append(nk)
            nn = ''
            if hasattr(data[nbr[0]], 'name') and nbr[1] < len(data[nbr[0]].name):
                nn = data[nbr[0]].name[nbr[1]]
            if is_anchor(nn):
                ah = nn[:40]
                break
    if ah:
        p['discovery_score'] = round(ds, 3)
        p['degree'] = d
        p['anchor'] = ah
        hidden.append(p)

hidden.sort(key=lambda p: -p['discovery_score'])
# Deduplicate
seen = {}
deduped = []
for p in hidden:
    k = p['name'].lower().strip()[:60]
    if k not in seen:
        seen[k] = p
        deduped.append(p)
    elif p['discovery_score'] > seen[k].get('discovery_score', 0):
        seen[k] = p
log(f"  {len(deduped)} unique hidden targets")

# ── PubMed check ──
log("\n[2/4] Checking PubMed (2022-2026) for all candidates...")
def check_pubmed(name):
    q = f'"{name}" AND (thrombosis OR coagulation OR platelet OR venous OR VTE)'
    params = {'db': 'pubmed', 'term': q, 'retmax': 1, 'retmode': 'xml',
              'mindate': '2022/01/01', 'maxdate': '2026/06/30', 'datetype': 'pdat'}
    url = f'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?{urllib.parse.urlencode(params)}'
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'VTE-GNN/1.0'})
        with urllib.request.urlopen(req, timeout=10) as r:
            root = ET.fromstring(r.read())
            return int(root.findtext('.//Count', '0'))
    except:
        return -1

for i, p in enumerate(deduped):
    if i % 50 == 0:
        log(f"  {i}/{len(deduped)} checked...")
    if i % 3 == 0:
        time.sleep(0.35)
    name_clean = p['name'].split(' (')[0].split(' -')[0].strip()[:60]
    p['pubmed_count'] = check_pubmed(name_clean)
    p['name_clean'] = name_clean

log(f"  Done. {sum(1 for p in deduped if p['pubmed_count']==0)} truly unknown (0 papers)")

# ── Write Excel ──
log("\n[3/4] Writing Excel workbook...")
wb = Workbook()

# Sheet 1: All hidden targets sorted by discovery score
ws1 = wb.active
ws1.title = "All Hidden Targets"
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

headers = ["Rank", "Target Name", "Type", "GNN Score", "Degree",
           "Discovery Score", "BFS Anchor", "PubMed 2022-2026", "Status"]
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

for i, p in enumerate(deduped):
    row = i + 2
    pc = p.get('pubmed_count', -1)
    if pc == 0:
        status = "TRULY UNKNOWN"
    elif pc > 0:
        status = "Validated"
    else:
        status = "N/A"

    ws1.cell(row=row, column=1, value=i+1)
    ws1.cell(row=row, column=2, value=p.get('name_clean', p['name'])[:80])
    ws1.cell(row=row, column=3, value=p['src_type'])
    ws1.cell(row=row, column=4, value=round(p['score'], 2))
    ws1.cell(row=row, column=5, value=p['degree'])
    ws1.cell(row=row, column=6, value=p['discovery_score'])
    ws1.cell(row=row, column=7, value=p.get('anchor', '')[:50])
    ws1.cell(row=row, column=8, value=pc if pc >= 0 else 'N/A')
    ws1.cell(row=row, column=9, value=status)

    if pc == 0:
        for col in range(1, 10):
            ws1.cell(row=row, column=col).fill = green_fill
    elif pc > 0 and pc <= 3:
        for col in range(1, 10):
            ws1.cell(row=row, column=col).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

ws1.auto_filter.ref = f"A1:I{len(deduped)+1}"
ws1.freeze_panes = "A2"

# Sheet 2: Truly Unknown only
ws2 = wb.create_sheet("Truly Unknown")
for col, h in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

unknown_only = [p for p in deduped if p.get('pubmed_count', -1) == 0]
unknown_only.sort(key=lambda p: -p['discovery_score'])
for i, p in enumerate(unknown_only):
    row = i + 2
    ws2.cell(row=row, column=1, value=i+1)
    ws2.cell(row=row, column=2, value=p.get('name_clean', p['name'])[:80])
    ws2.cell(row=row, column=3, value=p['src_type'])
    ws2.cell(row=row, column=4, value=round(p['score'], 2))
    ws2.cell(row=row, column=5, value=p['degree'])
    ws2.cell(row=row, column=6, value=p['discovery_score'])
    ws2.cell(row=row, column=7, value=p.get('anchor', '')[:50])
    ws2.cell(row=row, column=8, value=0)
    ws2.cell(row=row, column=9, value="TRULY UNKNOWN")

ws2.auto_filter.ref = f"A1:I{len(unknown_only)+1}"
ws2.freeze_panes = "A2"

# Sheet 3: Validated targets
ws3 = wb.create_sheet("Validated (>0 papers)")
for col, h in enumerate(headers, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

validated = [p for p in deduped if p.get('pubmed_count', -1) > 0]
validated.sort(key=lambda p: -p['discovery_score'])
for i, p in enumerate(validated):
    row = i + 2
    ws3.cell(row=row, column=1, value=i+1)
    ws3.cell(row=row, column=2, value=p.get('name_clean', p['name'])[:80])
    ws3.cell(row=row, column=3, value=p['src_type'])
    ws3.cell(row=row, column=4, value=round(p['score'], 2))
    ws3.cell(row=row, column=5, value=p['degree'])
    ws3.cell(row=row, column=6, value=p['discovery_score'])
    ws3.cell(row=row, column=7, value=p.get('anchor', '')[:50])
    ws3.cell(row=row, column=8, value=p.get('pubmed_count', 0))
    ws3.cell(row=row, column=9, value="Validated")

ws3.auto_filter.ref = f"A1:I{len(validated)+1}"
ws3.freeze_panes = "A2"

# Sheet 4: Summary stats
ws4 = wb.create_sheet("Summary")
ws4.cell(row=1, column=1, value="Hidden Target Hunt Summary").font = Font(bold=True, size=14)
ws4.cell(row=3, column=1, value="Model").font = Font(bold=True)
ws4.cell(row=3, column=2, value="PCA 128d (PubMedBERT + PCA reduction)")
ws4.cell(row=4, column=1, value="AUROC").font = Font(bold=True)
ws4.cell(row=4, column=2, value=0.925)
ws4.cell(row=5, column=1, value="MRR").font = Font(bold=True)
ws4.cell(row=5, column=2, value=0.232)
ws4.cell(row=6, column=1, value="Total candidates").font = Font(bold=True)
ws4.cell(row=6, column=2, value=len(deduped))
ws4.cell(row=7, column=1, value="Truly unknown (0 papers)").font = Font(bold=True)
ws4.cell(row=7, column=2, value=len(unknown_only))
ws4.cell(row=8, column=1, value="Validated (>0 papers)").font = Font(bold=True)
ws4.cell(row=8, column=2, value=len(validated))
ws4.cell(row=9, column=1, value="Validation rate").font = Font(bold=True)
ws4.cell(row=9, column=2, value=f"{100*len(validated)/max(len(deduped),1):.1f}%")
ws4.cell(row=10, column=1, value="PubMed search window").font = Font(bold=True)
ws4.cell(row=10, column=2, value="2022-01-01 to 2026-06-30")
ws4.cell(row=11, column=1, value="Generated").font = Font(bold=True)
ws4.cell(row=11, column=2, value=datetime.now().isoformat())

ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 40

# Save
output = r"D:\JY\work\my work\新思路\vte_gnn_target_discovery\figures\pca_hidden\hidden_targets_complete.xlsx"
os.makedirs(os.path.dirname(output), exist_ok=True)
wb.save(output)

t = time.time() - t0; m, s = divmod(t, 60)
log(f"\n[4/4] Saved: {output}")
log(f"Sheets: All ({len(deduped)}), Unknown ({len(unknown_only)}), Validated ({len(validated)}), Summary")
log(f"Done in {int(m)}m {int(s)}s")
